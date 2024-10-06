from openpyxl.styles import Font
from datetime import datetime
import openpyxl
import shutil
import os

class FixSpreadsheet:
    '''
    This class is used to fix the spreadsheets that are in the 'need_fixed' folder.
    '''

    def fix_data(self):
        '''
        This function is used to fix the data in the spreadsheets that are in the 'need_fixed' folder.
        '''
        directory = "spreadsheets/need_fixed"

        for file in os.listdir(directory):
            print(file)
            shutil.copyfile(f"spreadsheets/need_fixed/{file}", f"spreadsheets/fixed/{file}") # copy file from need_fixed to fixed

            path = f"spreadsheets/fixed/{file}"

            wb_obj = openpyxl.load_workbook(path) # load the IPU to be fixed
            sheet = wb_obj.active

            # get current value of GL string
            gl_string = sheet.cell(row=6, column=3).value
            gl_string = gl_string.replace("GL String:", "")

            for i in range(0, 3): # loop through and get rid of GL strings that don't match
                temp = sheet.cell(row=6+i, column=6).value
                if gl_string in temp:
                    continue
                else:
                    sheet.cell(row=6+i, column=6).value = ""

            # set date
            sheet.cell(row=16, column=1).value = "DATE: " + f"{(datetime.now()).month}/{(datetime.now()).day}/{(datetime.now()).year}"

            # switch sheet to 'Notes for Operations ONLY'
            for sheet2 in wb_obj:
                if "Notes" in sheet2.title:
                    sheet = sheet2

            # put licenses into a list
            licenses = sheet.cell(row=3, column=3).value
            licenses = licenses.replace("Original Lic #:", "")
            licenses = licenses.replace(" ", "")
            licenses = licenses.split(',')

            # rewrite licenses in a nicer format
            sheet.cell(row=3, column=3).value = "Original Lic #:"

            for i in range(0, len(licenses)):
                sheet.cell(row=4 + i, column=3).value = licenses[i] # put licenses under each other
                sheet.cell(row=4 + i, column=3).font = Font(name='Calibri', size=14, color="0070C0", bold=True)

            wb_obj.save(path)

if __name__ == "__main__":
    fixer = FixSpreadsheet()
    fixer.fix_data()