from datetime import datetime, timedelta
from asyncio.windows_events import NULL
from countryCodes import country_codes
from readSpreadsheet import DataHandler
from openpyxl.styles import Font
from nicknamer import NickName
import openpyxl
import logging
import shutil
import re

class WriteData:
    '''
    This class is responsible for writing the data from the Licensing spreadsheet and storing it in an IPU form.
    '''
    def __init__(self):
        self.nicknames = {}
        self.conflicts = set()
        pass

    def create_new_file(self, company, initials, dictionary):
        curr_time = datetime.now()
        formatted_name = company.replace("/", "").replace("\\", "") # if there are slashes in the company name, take them out so we can name the IPU file properly

        # name = input(f"Company name {company}, please enter a nickname: ") # ask for a company nickname from script runner to put on the IPU form
        name = ""
        for nickname in self.nicknames:
            if self.nicknames[nickname] == company:
                name = nickname
                break
        if name == "":
            while True:
                name = input(f"nickname not found for company {company}, please enter a nickname of less than 8 characters")
                if len(name) > 8:
                    print("nickname too long! please enter a nickname less than 8 characters")
                else:
                    break
            self.nicknames[name] = company
        
        default_email = "None"

        sub_list = [i for i in dictionary[company]['email'] if not isinstance(i, int)] # gets a list of only the valid emails
        if len(sub_list) > 0: # if there are valid emails, then put the IPU in the spreadsheets/completed/email folder
            shutil.copyfile("spreadsheets\IPU_form.xlsx", f"spreadsheets\completed\email\IPU-Clar2.0-{formatted_name}-{initials}.xlsx") # copy the IPU template, rename it, and put it in the proper folder
            path = f"spreadsheets\completed\email\IPU-Clar2.0-{formatted_name}-{initials}.xlsx" # change the file path to access the new blank IPU
            default_email = sub_list[0] # by default, the email that will be used is the first one read in from the company
        else: # if there are no emails found, copy the IPU template and rename it but put it in the 'without_email' folder
            shutil.copyfile("spreadsheets\IPU_form.xlsx", f"spreadsheets\completed\without_email\IPU-Clar2.0-{formatted_name}-{initials}.xlsx")
            path = f"spreadsheets\completed\without_email\IPU-Clar2.0-{formatted_name}-{initials}.xlsx"

        wb_obj = openpyxl.load_workbook(path) # load the blank IPU template
        sheet = wb_obj.active

        for clariti_sheet in wb_obj:
            if "Clariti" in clariti_sheet.title:
                sheet = clariti_sheet

        sheet.cell(row=3, column=2).value = "IPU-Clar2.0-" + name # change the IPU code to user provided company nickname

        # find and change the country code
        country_code = ""
        for code in country_codes: # this uses a dictionary of country codes, which you can find in 'countryCodes.py'
            if dictionary[company]['theater'].lower() == code: # if the country is in the dictionary, use the corresponding GL string
                country_code = country_codes[code]
        if country_code == "":
            logging.basicConfig(filename='./logs/vital_errors.log', encoding='utf-8',
                                level=logging.DEBUG)  # must have a 'logs' folder/directory in the project
            logging.info(f"Error finding country code for company {company}") # if it cannot find the country code, output an error
            sheet.cell(row=6, column=4).value = 0
        else:
            sheet.cell(row=6, column=4).value = country_code

        # make sure all country code fields are blank
        for i in range(0, 4):
            sheet.cell(row=5+i, column=6).value = ''


        # for each product license that is stored with the associated company
        for i in range(0, len(dictionary[company]['product id'])):
            count = i
            sheet.cell(row=19 + count, column=1).value = dictionary[company]['product id'][count] # product number (clarity 2.0 part)
            sheet.cell(row=19 + count, column=2).value = dictionary[company]['long description'][count] # product description
            sheet.cell(row=19 + count, column=3).value = dictionary[company]['quantity'][count] # quantity
            sheet.cell(row=19 + count, column=4).value = 0 # cost
            sheet.cell(row=19 + count, column=5).value = 0 # cost
            date = datetime.strftime(dictionary[company]['expiration date'][count], "%m/%d/%Y") # get expiration date
            # change term/dates so that its a week from today until expiration date
            week_out = curr_time+timedelta(days=7)
            formatted_date = f'{"0" * (2 - len(str(week_out.month))) + str(week_out.month)}/' \
                             f'{"0" * (2 - len(str(week_out.day))) + str(week_out.day)}/' \
                             f'{str(week_out.year)}'
            sheet.cell(row=19 + count, column=6).value = f"{formatted_date} to {date}"

        # customer information
        sheet.cell(row=36, column=2).value = company # full company name
        sheet.cell(row=37, column=2).value = dictionary[company]['address'] # address
        sheet.cell(row=39, column=2).value = f'{str(dictionary[company]["city"])} / {str(dictionary[company]["state"])} / {str(dictionary[company]["zip code"])}' # city / state / zipcode
        sheet.cell(row=40, column=2).value = dictionary[company]['country'] # country
        sheet.cell(row=41, column=2).value = dictionary[company]['contact name'] # primary contact name
        sheet.cell(row=43, column=2).value = default_email # email

        # adding the date edited
        formatted_date = f'{"0" * (2 - len(str(curr_time.month))) + str(curr_time.month)}/' \
                         f'{"0" * (2 - len(str(curr_time.day))) + str(curr_time.day)}/' \
                         f'{str(curr_time.year)}'
        sheet.cell(row=16, column=1).value = f'DATE: {formatted_date}'

        # switch sheet to 'Notes for Operations ONLY'
        for license_sheet in wb_obj:
            if "Notes" in license_sheet.title:
                sheet = license_sheet

        # append license numbers
        for i in range(0, len(dictionary[company]['license'])):
            sheet.cell(row=3 + i, column=4).value = dictionary[company]['license'][i]  # put licenses under each other
            sheet.cell(row=3 + i, column=4).font = Font(name='Calibri', size=18, color="0070C0", bold=True)

        wb_obj.save(path) # save the IPU

    def create_nicknames(self, dictionary):
        nicknamer = NickName()
        self.nicknames = nicknamer.nickname(dictionary)

    def mark_completed(self, companies):
        # only mark an IPO as completed if there are no conflicts

        path = f"spreadsheets\Licenses.xlsx"
        wb_obj = openpyxl.load_workbook(path) # load in the Licenses IPU template
        sheet = wb_obj.active

        m_row = 1000
        for i in range(2, m_row + 1):
            company = sheet.cell(column=2, row = i).value
            if company in companies and company not in self.conflicts: # if the company's IPO was processed and has no conflicts
                print(f"marking company {company} as complete...")
                nickname = "" # find the company's nickname
                nicknameFound = False
                for nn in self.nicknames:
                    if self.nicknames[nn] == company:
                        sheet.cell(column=5, row = i).value = "YES"
                        sheet.cell(column=6, row = i).value = f"IPU-Clar2.0-{nn}"
                        nicknameFound = True
                        break

                if not nicknameFound:
                    print(f"Error finding IPU code for company {company}, unable to mark as complete")       
            
        for company in self.conflicts:
            print(f"IPO information conflicts for company {company}, please review logs/conflicts.log before marking as completed")

        wb_obj.save(path)

    def process_files(self, initials):
        myData = DataHandler(initials=initials) # uses 'readSpreadsheet' to get your assigned IPU data
        data_dict = myData.get_data()
        self.conflicts = myData.check_validity() # checks to make sure emails are valid. If not, it will still create the IPUs but it will log conflicting
                                # emails to logs/conflicts.log
        self.create_nicknames(dictionary=data_dict)

        for company in data_dict.keys(): # for each company in the data dictionary
            self.create_new_file(company, initials=initials, dictionary=data_dict) # create a new IPU file

        self.mark_completed(companies=data_dict)

        
if __name__ == "__main__":
    makeFiles = WriteData()
    makeFiles.process_files(initials="LB") # put your initials here