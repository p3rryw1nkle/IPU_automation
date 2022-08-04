from datetime import datetime
from pprint import pprint
import openpyxl
import logging


class DataHandler:
    def __init__(self, initials):
        self.store_dict = {} # stores all IPU license information in a dictionary (data storage object)
        self.row_vals = []
        self.initials = initials # initials of the individual the IPU has been assigned to

    def get_data(self, path="spreadsheets\\Licenses.xlsx"):
        """
        Loads the data from each line in the spreadsheet
        :param path: where the licensing spreadsheet can be found. Default in spreadsheets folder.
        :return: a dictionary of companies, with the keys being company names.
        """
        # open the Licenses spreadsheet
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        # maximum number of rows to read through
        m_row = 1000
        for i in range(2, m_row + 1): # for each row in the 'Licenses' document

            initials_cell = sheet_obj.cell(row=i, column=4) # contains initials of who the IPU has been assigned to
            completed = sheet_obj.cell(row=i, column=5).value # contains whether or not the IPU has been completed or not

            # if the 'completed' cell has a value and its value is "y" or "yes" we skip the line
            if completed and completed.lower() in ["y", "yes"]:
                is_completed = True
            else:
                is_completed = False

            if initials_cell.value is not None and initials_cell.value.lower() == self.initials.lower() and not is_completed:
                # grab all the IPU information from the specific columns
                company_name = self.append_and_return(sheet_obj, row=i, col=2)
                license_num = self.append_and_return(sheet_obj, row=i, col=12)
                prod_id = self.append_and_return(sheet_obj, row=i, col=19)
                long_desc = self.append_and_return(sheet_obj, row=i, col=20)
                exp_date = self.append_and_return(sheet_obj, row=i, col=13)
                quantity = self.append_and_return(sheet_obj, row=i, col=10)
                phys_address = self.append_and_return(sheet_obj, row=i, col=21)
                city = self.append_and_return(sheet_obj, row=i, col=22)
                state = self.append_and_return(sheet_obj, row=i, col=23)
                zip_code = self.append_and_return(sheet_obj, row=i, col=25)
                email_address = self.append_and_return(sheet_obj, row=i, col=26)
                full_name = self.append_and_return(sheet_obj, row=i, col=27)
                country = self.append_and_return(sheet_obj, row=i, col=24)

                # if the company already has an entry in the dictionary, then there are multiple products/licenses.
                # So we are going to add the additional information to the company's dictionary.

                # formatting dict
                if company_name in self.store_dict:
                    self.store_dict[company_name]['license'].append(license_num)
                    self.store_dict[company_name]['product id'].append(prod_id)
                    self.store_dict[company_name]['long description'].append(long_desc)
                    self.store_dict[company_name]['expiration date'].append(exp_date)
                    self.store_dict[company_name]['quantity'].append(quantity)
                    self.store_dict[company_name]['email'].append(email_address if email_address not in self.store_dict[company_name]['email'] else 0)
                else: # if the company does not already have an entry in the IPU dictionary, create a new one
                    self.store_dict[company_name] = {'license': [license_num], # anything with square brackets around it [] there may be multiple of, so it's stored as an array
                                                     'product id': [prod_id],
                                                     'long description': [long_desc],
                                                     'expiration date': [exp_date],
                                                     'quantity': [quantity],
                                                     'address': phys_address,
                                                     'city': city,
                                                     'state': state,
                                                     'country': country,
                                                     'zip code': zip_code,
                                                     'email': [email_address],
                                                     'contact name': full_name
                                                     }

        return self.store_dict

    def append_and_return(self, sheet_obj, row, col):
        val = sheet_obj.cell(row=row, column=col).value
        self.row_vals.append(val)
        return val


    def check_validity(self):
        """
        Checks the validity of each email and full address stored for each company then logs which companies have errors
        :return: None
        """
        logging.basicConfig(filename='./logs/conflicts.log', encoding='utf-8', level=logging.DEBUG) # must have a 'logs' folder/directory in the project
        logging.info(f'Entry at {datetime.now()}')
        for company in self.store_dict.keys():
            sub_list = [i for i in self.store_dict[company]['email'] if not isinstance(i, int)] # for each company, take only valid emails (i.e. not Null, 0, etc.)
            if len(sub_list) > 1: # if there are more than 1 valid emails
                logging.info(f'Multiple emails for company: {company}. Emails found: {sub_list}') # list them in the logs/conflicts.log file

            # checking to see if the full address entered is valid
            if not isinstance(self.store_dict[company]['address'], str):
                logging.info(f'Invalid address for company: {company}. Address entered: {self.store_dict[company]["address"]}')  # list them in the logs/conflicts.log file

            if not isinstance(self.store_dict[company]['city'], str):
                logging.info(f'Invalid city for company: {company}. City entered: {self.store_dict[company]["city"]}')  # list them in the logs/conflicts.log file

            if not isinstance(self.store_dict[company]['state'], str):
                logging.info(f'Invalid state for company: {company}. State entered: {self.store_dict[company]["state"]}')  # list them in the logs/conflicts.log file

            if not isinstance(self.store_dict[company]['country'], str):
                logging.info(f'Invalid country for company: {company}. Country entered: {self.store_dict[company]["country"]}')  # list them in the logs/conflicts.log file


if __name__ == "__main__": # this code is only run if you run this script by itself, however the intention is to only run 'writeSpreadsheet'
    ss = DataHandler(initials="JR")

    dr = ss.get_data()
    pprint(dr)
    ss.check_validity()
